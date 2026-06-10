"""
news_history.py — News History Module
Refactored to save fetched news to SQLite database instead of writing to static Excel files.
Supports dynamic generation of Excel files on demand to preserve API compatibility.
"""
import os
from datetime import datetime, timezone
from typing import List, Set, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import NewsItem
from config import settings
from database.db import SessionLocal
import database.crud as db_crud

# Column definitions for Excel generation
COLUMNS = [
    ("News Heading",     "title",           60),
    ("News Link",        "url",             50),
    ("Sentiment Effect", "sentiment",       18),
    ("Sentiment Score",  "sentiment_score", 16),
    ("Source",           "source",          20),
    ("Published Date",   "published",       22),
    ("Saved At",         "saved_at",        22),
]

HEADER_FILL   = PatternFill("solid", fgColor="1E293B")   # dark slate
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri")
BULLISH_FILL  = PatternFill("solid", fgColor="D1FAE5")   # light green
BEARISH_FILL  = PatternFill("solid", fgColor="FEE2E2")   # light red
NEUTRAL_FILL  = PatternFill("solid", fgColor="F1F5F9")   # light gray


def get_history_path(category: str = "General") -> str:
    """
    Dynamically generates a styled Excel file from database records.
    Returns the file path for the web server to send as response.
    """
    # Create temp directory
    temp_dir = settings.DATA_DIR / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    excel_file = temp_dir / f"{category.lower().replace(' ', '_')}_news_history.xlsx"

    with SessionLocal() as db:
        articles = db_crud.get_news_articles(db, category, limit=1000)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{category} News"

    # Style Header Row
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Write Data
    for row_idx, art in enumerate(articles, start=2):
        sentiment = art.sentiment.upper()
        fill = BULLISH_FILL if sentiment == "BULLISH" else (BEARISH_FILL if sentiment == "BEARISH" else NEUTRAL_FILL)

        pub_str = art.published.strftime("%Y-%m-%d %H:%M:%S") if isinstance(art.published, datetime) else str(art.published)
        created_str = art.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(art.created_at, datetime) else str(art.created_at)

        row_data = [
            art.headline,
            art.url,
            sentiment,
            round(art.impact_score, 4),
            art.source,
            pub_str,
            created_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if col_idx == 2 and art.url:
                cell.hyperlink = art.url
                cell.font = Font(color="2563EB", underline="single", name="Calibri")
            else:
                cell.font = Font(name="Calibri")

    wb.save(str(excel_file))
    return str(excel_file)


def save_news_to_excel(news_items: List[NewsItem]) -> int:
    """
    Saves news items directly to database 'news_articles' table.
    Returns the total count of new records successfully added.
    """
    if not news_items:
        return 0

    news_dicts = []
    for item in news_items:
        sentiment_val = getattr(item, "sentiment", "NEUTRAL")
        if hasattr(sentiment_val, "value"):
            sentiment_val = sentiment_val.value
        sentiment_val = str(sentiment_val).upper().replace("SENTIMENTTYPE.", "")

        news_dicts.append({
            "headline": item.title.strip(),
            "summary": "",
            "source": item.source.strip(),
            "url": item.url.strip(),
            "published": item.published,
            "sentiment": sentiment_val,
            "impact_score": item.sentiment_score,
            "category": getattr(item, "category", "General"),
            "related_symbols": getattr(item, "related_symbols", []),
        })

    with SessionLocal() as db:
        added_count = db_crud.create_news_articles_batch(db, news_dicts)
    
    if added_count > 0:
        print(f"[news_history] Saved {added_count} new news items to database")
    else:
        print("[news_history] No new news items (all duplicates in database).")
        
    return added_count


def get_history_stats(category: str = "General") -> dict:
    """Return stats about the database news records for category."""
    with SessionLocal() as db:
        from database.models import NewsArticle
        total = db.query(NewsArticle).filter(NewsArticle.category == category).count()
    return {
        "total_rows": total,
        "file_exists": True,
        "file_path": f"Database (Category: {category})"
    }


def get_all_history_stats() -> dict:
    categories = ["Indian Markets", "Crypto", "Commodities", "Global Markets", "General"]
    stats = {}
    for cat in categories:
        stats[cat] = get_history_stats(cat)
    return stats


def reset_history(category: str) -> bool:
    """Delete all news records in database for this category."""
    with SessionLocal() as db:
        return db_crud.delete_news_by_category(db, category)

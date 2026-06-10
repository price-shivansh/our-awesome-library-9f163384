"""
migrate_data.py — Data Migration script.
Imports JSON files, SQLite databases, and Excel sheets into the new quant.db unified schema.
Optimized to handle duplicate records in Excel spreadsheets without transaction crashes.
"""
import os
import sys
import json
import sqlite3
import glob
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

# Adjust python path to find database modules
sys.path.append(str(Path(__file__).resolve().parent.parent))

from database.db import Base, engine, SessionLocal
from database.models import (
    TelegramSubscriber,
    SentNews,
    PredictionMemory,
    PredictionOutcome,
    SetupStats,
    AdaptiveWeight,
    WeightHistory,
    NewsArticle,
)

def parse_date(date_str) -> datetime:
    """Parse date from string with fallbacks."""
    if not date_str:
        return datetime.now(timezone.utc)
    if isinstance(date_str, datetime):
        return date_str
    
    # Clean string
    date_str = str(date_str).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        from dateutil import parser
        return parser.parse(date_str).replace(tzinfo=timezone.utc)
    except Exception:
        return datetime.now(timezone.utc)

def make_naive(dt) -> datetime:
    """Strip timezone info for standard database comparison."""
    if isinstance(dt, datetime):
        return dt.replace(tzinfo=None)
    return dt

def find_file(data_dir, filename) -> Optional[Path]:
    """Find file in data_dir or backup/."""
    p1 = data_dir / filename
    if p1.exists():
        return p1
    p2 = data_dir / "backup" / filename
    if p2.exists():
        return p2
    return None

def migrate_json_subscribers(data_dir):
    json_path = find_file(data_dir, "telegram_subscribers.json")
    if not json_path:
        print("[Migration] JSON Subscribers file not found. Skipping.")
        return 0

    print(f"[Migration] Found subscribers at {json_path}. Migrating...")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    subs = data.get("subscribers", {})
    count = 0
    with SessionLocal() as db:
        for chat_id, prefs in subs.items():
            exists = db.query(TelegramSubscriber).filter_by(chat_id=str(chat_id)).first()
            if exists:
                continue
            db_sub = TelegramSubscriber(
                chat_id=str(chat_id),
                is_active=prefs.get("is_active", True),
                filters=json.dumps(prefs.get("filters", {}))
            )
            db.add(db_sub)
            count += 1
        db.commit()
    print(f"[Migration] Successfully migrated {count} subscribers.")
    return count

def migrate_json_sent_news(data_dir):
    json_path = find_file(data_dir, "sent_news.json")
    if not json_path:
        print("[Migration] JSON Sent News file not found. Skipping.")
        return 0

    print(f"[Migration] Found sent news at {json_path}. Migrating...")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    headlines_dict = {}
    if isinstance(data, list):
        old_ts = datetime.now(timezone.utc)
        headlines_dict = {title: old_ts for title in data}
    elif isinstance(data, dict):
        headlines_dict = data

    count = 0
    with SessionLocal() as db:
        for headline, ts_str in headlines_dict.items():
            exists = db.query(SentNews).filter_by(headline=headline).first()
            if exists:
                continue
            db_item = SentNews(
                headline=headline,
                sent_at=parse_date(ts_str)
            )
            db.add(db_item)
            count += 1
        db.commit()
    print(f"[Migration] Successfully migrated {count} sent headlines.")
    return count

def migrate_sqlite_memory(data_dir):
    db_path = find_file(data_dir, "adaptive_memory.db")
    if not db_path:
        print("[Migration] SQLite database adaptive_memory.db not found. Skipping.")
        return

    print(f"[Migration] Found SQLite memory at {db_path}. Connecting...")
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # 1. Migrate Adaptive Weights
    try:
        cursor.execute("SELECT * FROM adaptive_weights")
        weights = cursor.fetchall()
        w_count = 0
        with SessionLocal() as db:
            for w_row in weights:
                w = dict(w_row)
                exists = db.query(AdaptiveWeight).filter_by(
                    weight_key=w["weight_key"], symbol=w["symbol"]
                ).first()
                if exists:
                    continue
                db_weight = AdaptiveWeight(
                    weight_key=w["weight_key"],
                    symbol=w["symbol"],
                    value=w["value"],
                    default_value=w["default_value"],
                    last_updated=parse_date(w["last_updated"]),
                )
                db.add(db_weight)
                w_count += 1
            db.commit()
        print(f"[Migration] Migrated {w_count} adaptive weights.")
    except Exception as e:
        print(f"[Migration] Error migrating weights: {e}")

    # 2. Migrate Weight History
    try:
        cursor.execute("SELECT * FROM weight_history")
        history = cursor.fetchall()
        h_count = 0
        with SessionLocal() as db:
            for h_row in history:
                h = dict(h_row)
                db_hist = WeightHistory(
                    weight_key=h["weight_key"],
                    symbol=h["symbol"],
                    old_value=h["old_value"],
                    new_value=h["new_value"],
                    reason=h["reason"],
                    changed_at=parse_date(h["changed_at"]),
                )
                db.add(db_hist)
                h_count += 1
            db.commit()
        print(f"[Migration] Migrated {h_count} weight history records.")
    except Exception as e:
        print(f"[Migration] Error migrating weight history: {e}")

    # 3. Migrate Setup Stats
    try:
        cursor.execute("SELECT * FROM setup_stats")
        stats = cursor.fetchall()
        s_count = 0
        with SessionLocal() as db:
            for s_row in stats:
                s = dict(s_row)
                exists = db.query(SetupStats).filter_by(
                    setup_name=s["setup_name"], symbol=s["symbol"]
                ).first()
                if exists:
                    continue
                db_stats = SetupStats(
                    setup_name=s["setup_name"],
                    symbol=s["symbol"],
                    total_predictions=s["total_predictions"],
                    correct_count=s["correct_count"],
                    incorrect_count=s["incorrect_count"],
                    neutral_count=s.get("neutral_count", 0),
                    win_rate=s["win_rate"],
                    avg_return_pct=s["avg_return_pct"],
                    last_updated=parse_date(s["last_updated"]),
                )
                db.add(db_stats)
                s_count += 1
            db.commit()
        print(f"[Migration] Migrated {s_count} setup stats records.")
    except Exception as e:
        print(f"[Migration] Error migrating setup stats: {e}")

    # 4. Migrate Predictions and map IDs for Outcomes
    pred_id_mapping = {}  # old_id -> new_id
    try:
        cursor.execute("SELECT * FROM predictions")
        predictions = cursor.fetchall()
        p_count = 0
        with SessionLocal() as db:
            for p_row in predictions:
                p = dict(p_row)
                created_dt = parse_date(p["timestamp"])
                exists = db.query(PredictionMemory).filter_by(
                    symbol=p["symbol"], created_at=created_dt
                ).first()
                
                if exists:
                    pred_id_mapping[p["id"]] = exists.id
                    continue

                db_pred = PredictionMemory(
                    symbol=p["symbol"],
                    timeframe="1h",
                    prediction=p["bias"],
                    confidence=p["confidence_score"],
                    technical_score=p["technical_score"],
                    sentiment_score=p["news_score"],
                    market_regime=p.get("market_regime"),
                    created_at=created_dt,
                    rsi_value=p.get("rsi_value"),
                    macd_signal=p.get("macd_signal"),
                    ema_trend=p.get("ema_trend"),
                    momentum_signal=p.get("momentum_signal"),
                    bb_signal=p.get("bb_signal"),
                    atr_pct=p.get("atr_pct"),
                    price_at_prediction=p["price_at_prediction"],
                    active_setups=p.get("active_setups", "[]"),
                )
                db.add(db_pred)
                db.flush()  # Populate id
                pred_id_mapping[p["id"]] = db_pred.id
                p_count += 1
            db.commit()
        print(f"[Migration] Migrated {p_count} predictions.")
    except Exception as e:
        print(f"[Migration] Error migrating predictions: {e}")

    # 5. Migrate Outcomes
    try:
        cursor.execute("SELECT * FROM outcomes")
        outcomes = cursor.fetchall()
        o_count = 0
        with SessionLocal() as db:
            for o_row in outcomes:
                o = dict(o_row)
                old_pred_id = o["prediction_id"]
                new_pred_id = pred_id_mapping.get(old_pred_id)
                if not new_pred_id:
                    continue

                exists = db.query(PredictionOutcome).filter_by(
                    prediction_id=new_pred_id, horizon=o["horizon"]
                ).first()
                if exists:
                    continue

                db_outcome = PredictionOutcome(
                    prediction_id=new_pred_id,
                    horizon=o["horizon"],
                    price_at_outcome=o["price_at_outcome"],
                    price_change=o["price_change_pct"],
                    outcome=o["outcome"],
                    success=(o["outcome"] == "CORRECT"),
                    actual_direction="Neutral",
                    evaluated_at=parse_date(o["evaluated_at"]),
                )
                db.add(db_outcome)
                o_count += 1
            db.commit()
        print(f"[Migration] Migrated {o_count} outcomes.")
    except Exception as e:
        print(f"[Migration] Error migrating outcomes: {e}")

    conn.close()

def migrate_historical_spreadsheets(backend_dir, data_dir):
    from openpyxl import load_workbook
    
    # 1. Populate seen sets from DB to avoid UNIQUE constraint violations
    seen_urls = set()
    seen_headlines = set()
    with SessionLocal() as db:
        rows = db.query(NewsArticle.url, NewsArticle.headline, NewsArticle.source, NewsArticle.published).all()
        for r in rows:
            if r.url:
                seen_urls.add(r.url.strip())
            seen_headlines.add((r.headline.strip(), r.source.strip(), make_naive(r.published)))

    # 2. Migrate Category Sheets
    history_dir = backend_dir / "History"
    category_files = {
        "commodities_news_history.xlsx": "Commodities",
        "crypto_news_history.xlsx": "Crypto",
        "global_markets_news_history.xlsx": "Global Markets",
        "indian_markets_news_history.xlsx": "Indian Markets",
    }

    n_count = 0
    for filename, category in category_files.items():
        filepath = history_dir / filename
        if not filepath.exists():
            print(f"[Migration] Category sheet not found at {filepath}. Skipping.")
            continue

        print(f"[Migration] Parsing {filepath} ({category})...")
        try:
            wb = load_workbook(str(filepath), read_only=True)
            ws = wb.active
            rows_added = 0
            
            with SessionLocal() as db:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    title = str(row[0]).strip()
                    url = str(row[1] or "").strip()
                    sentiment = str(row[2]).upper().strip()
                    sentiment_score = float(row[3]) if row[3] is not None else 0.0
                    source = str(row[4]).strip()
                    published = parse_date(row[5])
                    first_seen = parse_date(row[6]) if len(row) > 6 else published
                    
                    unique_key = (title, source, make_naive(published))
                    if not url or url in seen_urls or unique_key in seen_headlines:
                        continue
                    seen_urls.add(url)
                    seen_headlines.add(unique_key)

                    db_art = NewsArticle(
                        headline=title,
                        summary="",
                        source=source,
                        url=url,
                        published=published,
                        sentiment=sentiment,
                        impact_score=sentiment_score,
                        first_seen_at=first_seen,
                        category=category,
                        related_symbols=None,
                    )
                    db.add(db_art)
                    rows_added += 1
                    n_count += 1
                db.commit()
            wb.close()
            print(f"[Migration] Migrated {rows_added} news articles from {filename}.")
        except Exception as e:
            print(f"[Migration] Error parsing spreadsheet {filename}: {e}")

    # 3. Migrate Daily Archives
    archive_dir = data_dir / "news_archive"
    if archive_dir.exists():
        archive_files = glob.glob(str(archive_dir / "news_*.xlsx"))
        print(f"[Migration] Found {len(archive_files)} daily news spreadsheets in archive. Migrating...")
        
        for filepath in archive_files:
            try:
                wb = load_workbook(filepath, read_only=True)
                ws = wb.active
                rows_added = 0
                
                headers = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = list(row) if row else []
                    break
                
                if not headers:
                    wb.close()
                    continue
                    
                col_map = {name: idx for idx, name in enumerate(headers) if name is not None}
                
                with SessionLocal() as db:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or not any(x is not None for x in row):
                            continue
                            
                        def get_val(col_name, default=""):
                            idx = col_map.get(col_name)
                            if idx is not None and idx < len(row):
                                val = row[idx]
                                return val if val is not None else default
                            return default
                        
                        title = str(get_val("title")).strip()
                        url = str(get_val("url")).strip()
                        source = str(get_val("source")).strip()
                        published = parse_date(get_val("timestamp"))
                        sentiment = str(get_val("sentiment")).upper().strip()
                        sentiment_score = float(get_val("sentiment_score", 0.0))
                        category = str(get_val("category", "General")).strip()
                        related = str(get_val("related_symbols", ""))
                        
                        unique_key = (title, source, make_naive(published))
                        if not url or url in seen_urls or unique_key in seen_headlines:
                            continue
                        seen_urls.add(url)
                        seen_headlines.add(unique_key)
                            
                        db_art = NewsArticle(
                            headline=title,
                            summary="",
                            source=source,
                            url=url,
                            published=published,
                            sentiment=sentiment,
                            impact_score=sentiment_score,
                            first_seen_at=published,
                            category=category,
                            related_symbols=related if related else None,
                        )
                        db.add(db_art)
                        rows_added += 1
                        n_count += 1
                    db.commit()
                wb.close()
            except Exception as e:
                print(f"[Migration] Error parsing daily archive {os.path.basename(filepath)}: {e}")
                
    print(f"[Migration] Total daily archive news articles migrated: {n_count}")

def backup_original_files(root_dir, data_dir):
    print("[Migration] Archiving original files to .backup...")
    backup_dir = data_dir / "backup"
    backup_dir.mkdir(parents=True, exist_ok=True)

    # Backup JSONs
    for filename in ("telegram_subscribers.json", "sent_news.json"):
        source = data_dir / filename
        if source.exists():
            dest = backup_dir / filename
            try:
                if dest.exists():
                    dest.unlink()
                source.rename(dest)
                print(f"[Migration] Archived JSON {filename} to {dest}")
            except Exception as e:
                print(f"[Migration] Failed to backup JSON {filename}: {e}")

    # Backup sqlite db
    db_source = data_dir / "adaptive_memory.db"
    if db_source.exists():
        db_dest = backup_dir / "adaptive_memory.db"
        try:
            if db_dest.exists():
                db_dest.unlink()
            db_source.rename(db_dest)
            print(f"[Migration] Archived Database adaptive_memory.db to {db_dest}")
        except Exception as e:
            print(f"[Migration] Failed to backup Database: {e}")
            
    # Clean WAL & SHM files
    for ext in ("-shm", "-wal"):
        file_path = data_dir / f"adaptive_memory.db{ext}"
        if file_path.exists():
            try:
                file_path.unlink()
            except Exception:
                pass

def main():
    backend_dir = Path(__file__).resolve().parent.parent
    root_dir = backend_dir.parent
    data_dir = backend_dir / "data"

    print("====================================================")
    print("           Quant Decision Engine Migration          ")
    print("====================================================")
    
    # Enable schema check/creation
    Base.metadata.create_all(bind=engine)
    
    try:
        migrate_json_subscribers(data_dir)
        migrate_json_sent_news(data_dir)
        migrate_sqlite_memory(data_dir)
        migrate_historical_spreadsheets(backend_dir, data_dir)
        
        # Verify counts
        print("----------------------------------------------------")
        print("DB Row Counts Verification:")
        with SessionLocal() as db:
            print(f"  NewsArticles:         {db.query(NewsArticle).count()}")
            print(f"  PredictionMemory:    {db.query(PredictionMemory).count()}")
            print(f"  PredictionOutcome:   {db.query(PredictionOutcome).count()}")
            print(f"  SetupStats:          {db.query(SetupStats).count()}")
            print(f"  AdaptiveWeight:      {db.query(AdaptiveWeight).count()}")
            print(f"  WeightHistory:       {db.query(WeightHistory).count()}")
            print(f"  TelegramSubscriber:  {db.query(TelegramSubscriber).count()}")
            print(f"  SentNewsCache:       {db.query(SentNews).count()}")
        print("----------------------------------------------------")
        
        # Safe archival
        backup_original_files(root_dir, data_dir)
        print("[Migration] SUCCESS. Data migration completed.")
    except Exception as e:
        print(f"[Migration] FATAL ERROR during migration: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

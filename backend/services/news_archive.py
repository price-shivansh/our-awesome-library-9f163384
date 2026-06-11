"""
services/news_archive.py — Unified News Archiving & History Service

Merged from:
  - services/news_archive.py  (DbNewsArchive class — date-based query interface)
  - news_history.py           (Excel export, stats, save helpers)

All consumers should import from here. The old news_history.py has been removed.
"""
from typing import List, Dict, Optional, Set
from datetime import datetime, timezone
from pathlib import Path

from database.db import SessionLocal
import database.crud as db_crud

try:
    from schemas.market_schemas import NewsItem
except ImportError:
    NewsItem = None  # type: ignore[assignment,misc]


# ── Excel styling constants (used for on-demand Excel export) ─────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

COLUMNS = [
    ("News Heading",     "title",           60),
    ("News Link",        "url",             50),
    ("Sentiment Effect", "sentiment",       18),
    ("Sentiment Score",  "sentiment_score", 16),
    ("Source",           "source",          20),
    ("Published Date",   "published",       22),
    ("Saved At",         "saved_at",        22),
]

HEADER_FILL   = None
HEADER_FONT   = None
BULLISH_FILL  = None
BEARISH_FILL  = None
NEUTRAL_FILL  = None

if _OPENPYXL_AVAILABLE:
    HEADER_FILL   = PatternFill("solid", fgColor="1E293B")   # dark slate
    HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri")
    BULLISH_FILL  = PatternFill("solid", fgColor="D1FAE5")   # light green
    BEARISH_FILL  = PatternFill("solid", fgColor="FEE2E2")   # light red
    NEUTRAL_FILL  = PatternFill("solid", fgColor="F1F5F9")   # light gray


# ── Internal helpers ──────────────────────────────────────────────────────────

def _normalise_sentiment(sentiment_val) -> str:
    """Normalise a sentiment value (enum or string) to an uppercase plain string."""
    if hasattr(sentiment_val, "value"):
        sentiment_val = sentiment_val.value
    return str(sentiment_val).upper().replace("SENTIMENTTYPE.", "")


def _news_item_to_dict(item) -> Optional[dict]:
    """Convert a NewsItem object to a dict suitable for db_crud.create_news_articles_batch."""
    if not item:
        return None
    published = getattr(item, "published", None)
    if not published:
        return None
    return {
        "headline": item.title.strip(),
        "summary": "",
        "source": item.source.strip(),
        "url": item.url.strip(),
        "published": published,
        "sentiment": _normalise_sentiment(getattr(item, "sentiment", "NEUTRAL")),
        "impact_score": float(getattr(item, "sentiment_score", 0.0)),
        "category": getattr(item, "category", "General"),
        "related_symbols": getattr(item, "related_symbols", []),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Abstract Base
# ═══════════════════════════════════════════════════════════════════════════════

class BaseNewsArchive:
    """Abstract base for news archiving system. Allows swapping storage backends."""

    def archive_news(self, news_items: list) -> int:
        raise NotImplementedError

    def get_by_date(self, date_str: str) -> List[Dict]:
        """date_str in YYYY-MM-DD format"""
        raise NotImplementedError

    def get_by_date_range(self, start_date: str, end_date: str) -> List[Dict]:
        """start_date, end_date in YYYY-MM-DD format"""
        raise NotImplementedError

    def get_available_dates(self) -> List[str]:
        raise NotImplementedError


# ═══════════════════════════════════════════════════════════════════════════════
# SQLite Implementation
# ═══════════════════════════════════════════════════════════════════════════════

class DbNewsArchive(BaseNewsArchive):
    """
    SQLAlchemy-backed implementation of the news archiving system.
    Stores and queries all archived items in the unified quant.db database.
    """

    # ── Write operations ───────────────────────────────────────────────────────

    def archive_news(self, news_items: list) -> int:
        """
        Save a list of NewsItem objects to the database.
        Returns the number of newly inserted records (duplicates skipped).
        """
        if not news_items:
            return 0
        news_dicts = [d for d in (_news_item_to_dict(item) for item in news_items) if d]
        if not news_dicts:
            return 0
        with SessionLocal() as db:
            added = db_crud.create_news_articles_batch(db, news_dicts)
        if added > 0:
            print(f"[news_archive] Saved {added} new news items to database")
        else:
            print("[news_archive] No new news items (all duplicates in database).")
        return added

    # Alias for backward compatibility with consumers that called save_news_to_excel()
    def save_news(self, news_items: list) -> int:
        """Alias for archive_news() — replaces news_history.save_news_to_excel()."""
        return self.archive_news(news_items)

    # ── Read operations ────────────────────────────────────────────────────────

    def get_by_date(self, date_str: str) -> List[Dict]:
        with SessionLocal() as db:
            articles = db_crud.get_news_by_date(db, date_str)
        return [self._article_to_dict(art) for art in articles]

    def get_by_date_range(self, start_date: str, end_date: str) -> List[Dict]:
        with SessionLocal() as db:
            articles = db_crud.get_news_by_date_range(db, start_date, end_date)
        return [self._article_to_dict(art) for art in articles]

    def get_available_dates(self) -> List[str]:
        with SessionLocal() as db:
            return db_crud.get_news_available_dates(db)

    @staticmethod
    def _article_to_dict(art) -> dict:
        related = art.related_symbols.split(",") if art.related_symbols else []
        pub_str = (
            art.published.strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(art.published, datetime)
            else str(art.published)
        )
        return {
            "title": art.headline,
            "source": art.source,
            "url": art.url,
            "published": pub_str,
            "sentiment": art.sentiment,
            "sentiment_score": float(art.impact_score),
            "category": art.category,
            "related_symbols": related,
        }

    # ── Excel export (on-demand, replaces news_history.get_history_path) ──────

    def get_history_path(self, category: str = "General") -> str:
        """
        Dynamically generate a styled Excel file from database records.
        Returns the file path for the web server to send as a response.
        Replaces: news_history.get_history_path()
        """
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed; cannot generate Excel file.")

        from config import settings  # imported here to avoid circular issues

        temp_dir = Path(settings.DATA_DIR) / "temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        excel_file = temp_dir / f"{category.lower().replace(' ', '_')}_news_history.xlsx"

        with SessionLocal() as db:
            articles = db_crud.get_news_articles(db, category, limit=1000)

        wb = Workbook()
        ws = wb.active
        ws.title = f"{category} News"

        # Header row
        for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Data rows
        for row_idx, art in enumerate(articles, start=2):
            sentiment = art.sentiment.upper()
            fill = (
                BULLISH_FILL if sentiment == "BULLISH"
                else (BEARISH_FILL if sentiment == "BEARISH" else NEUTRAL_FILL)
            )
            pub_str = (
                art.published.strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(art.published, datetime) else str(art.published)
            )
            created_str = (
                art.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(art.created_at, datetime) else str(art.created_at)
            )
            row_data = [
                art.headline, art.url, sentiment,
                round(art.impact_score, 4), art.source, pub_str, created_str,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=False, vertical="center")
                if col_idx == 2 and art.url:
                    cell.hyperlink = art.url
                    cell.font = Font(color="2563EB", underline="single", name="Calibri")
                else:
                    cell.font = Font(name="Calibri")

        wb.save(str(excel_file))
        return str(excel_file)

    # ── Statistics helpers (replaces news_history.get_history_stats) ──────────

    def get_history_stats(self, category: str = "General") -> dict:
        """Return stats about DB news records for a category."""
        from database.models import NewsArticle
        with SessionLocal() as db:
            total = db.query(NewsArticle).filter(NewsArticle.category == category).count()
        return {
            "total_rows": total,
            "file_exists": True,
            "file_path": f"Database (Category: {category})",
        }

    def get_all_history_stats(self) -> dict:
        """Return stats for all standard categories."""
        categories = ["Indian Markets", "Crypto", "Commodities", "Global Markets", "General"]
        return {cat: self.get_history_stats(cat) for cat in categories}

    def reset_history(self, category: str) -> bool:
        """Delete all news records in database for this category."""
        with SessionLocal() as db:
            return db_crud.delete_news_by_category(db, category)


# ── Singleton instance (used by all consumers) ─────────────────────────────────
news_archive_service = DbNewsArchive()


# ── Module-level convenience shims (for routers/sentiment.py backward compat) ──

def get_history_path(category: str = "General") -> str:
    """Module-level shim — delegates to news_archive_service.get_history_path()."""
    return news_archive_service.get_history_path(category)


def save_news_to_excel(news_items: list) -> int:
    """Module-level shim — delegates to news_archive_service.archive_news()."""
    return news_archive_service.archive_news(news_items)


def get_history_stats(category: str = "General") -> dict:
    """Module-level shim — delegates to news_archive_service.get_history_stats()."""
    return news_archive_service.get_history_stats(category)


def get_all_history_stats() -> dict:
    """Module-level shim — delegates to news_archive_service.get_all_history_stats()."""
    return news_archive_service.get_all_history_stats()


def reset_history(category: str) -> bool:
    """Module-level shim — delegates to news_archive_service.reset_history()."""
    return news_archive_service.reset_history(category)
